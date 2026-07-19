"""Document parsers: PDF / DOCX / XLSX / Markdown -> plain text.

Each parser is a pure function ``(Path) -> str``. Output is normalised to a
lightweight Markdown-ish form so the downstream chunker can be *heading-aware*
regardless of source format:

* Markdown headings (``# ...``) are preserved verbatim.
* DOCX ``Heading N`` paragraph styles are re-emitted as ``#`` * N headings.
* XLSX sheet names become ``#`` headings; rows become tab-joined lines.
* PDF has no reliable heading structure, so pages are emitted as plain text
  separated by blank lines (the chunker still windows them).

No network, no cloud services — everything runs on locally-installed libs.
"""

from __future__ import annotations

from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

# Extensions we know how to parse, mapped by the dispatcher below.
SUPPORTED_SUFFIXES = frozenset({".pdf", ".docx", ".xlsx", ".md", ".markdown", ".txt"})


class UnsupportedFileError(ValueError):
    """Raised when a file's extension has no registered parser."""


def parse_pdf(path: Path) -> str:
    """Extract text from a PDF, one blank line between pages."""
    reader = PdfReader(str(path))
    pages = [(page.extract_text() or "").strip() for page in reader.pages]
    return "\n\n".join(p for p in pages if p)


def parse_docx(path: Path) -> str:
    """Extract text from a DOCX, mapping ``Heading N`` styles to Markdown."""
    document = Document(str(path))
    lines: list[str] = []
    for para in document.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        style = (para.style.name or "") if para.style else ""
        if style.startswith("Heading"):
            # "Heading 2" -> level 2; "Heading" / "Title" -> level 1.
            tail = style.replace("Heading", "").strip()
            level = int(tail) if tail.isdigit() else 1
            lines.append(f"{'#' * min(level, 6)} {text}")
        else:
            lines.append(text)
    return "\n\n".join(lines)


def parse_xlsx(path: Path) -> str:
    """Extract cell text from an XLSX; each sheet becomes a Markdown heading."""
    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    blocks: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                rows.append("\t".join(cells))
        if rows:
            blocks.append(f"# {sheet.title}\n" + "\n".join(rows))
    workbook.close()
    return "\n\n".join(blocks)


def parse_markdown(path: Path) -> str:
    """Read a Markdown / plain-text file as-is (UTF-8)."""
    return path.read_text(encoding="utf-8")


def parse_file(path: str | Path) -> str:
    """Dispatch to the parser for ``path``'s extension.

    Args:
        path: File to parse.

    Returns:
        Normalised plain text (see module docstring).

    Raises:
        UnsupportedFileError: If the extension has no registered parser.
        FileNotFoundError: If ``path`` does not exist.
    """
    p = Path(path)
    if not p.is_file():
        raise FileNotFoundError(f"No such file: {p}")
    suffix = p.suffix.lower()
    if suffix == ".pdf":
        return parse_pdf(p)
    if suffix == ".docx":
        return parse_docx(p)
    if suffix == ".xlsx":
        return parse_xlsx(p)
    if suffix in (".md", ".markdown", ".txt"):
        return parse_markdown(p)
    raise UnsupportedFileError(
        f"Cannot parse {p.name!r}: supported extensions are "
        f"{', '.join(sorted(SUPPORTED_SUFFIXES))}"
    )
